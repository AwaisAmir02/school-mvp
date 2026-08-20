from dataclasses import dataclass, field

import openpyxl
from django.db import IntegrityError, transaction

from academics.models import ClassLevel, Section
from students.models import Student

REQUIRED_COLUMNS = ['admission_number', 'first_name', 'last_name', 'class_level']
EXPECTED_COLUMNS = [
    'admission_number',
    'first_name',
    'last_name',
    'date_of_birth',
    'gender',
    'class_level',
    'section',
    'guardian_name',
    'guardian_phone',
    'email',
]


class StudentImportError(Exception):
    pass


@dataclass
class RowResult:
    row_number: int
    admission_number: str
    status: str
    errors: list = field(default_factory=list)


@dataclass
class ImportSummary:
    total_rows: int = 0
    created_count: int = 0
    error_count: int = 0
    row_results: list = field(default_factory=list)


def _read_header(sheet):
    header = []
    for cell in sheet[1]:
        header.append(str(cell.value).strip().lower() if cell.value is not None else '')
    return header


def _row_values(row, column_index):
    return {name: (row[idx].value if idx < len(row) else None) for name, idx in column_index.items()}


def import_students(*, file, school, academic_year):
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active
    header = _read_header(sheet)
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in header]
    if missing_columns:
        raise StudentImportError(f"Missing required columns: {', '.join(missing_columns)}")

    column_index = {name: idx for idx, name in enumerate(header) if name in EXPECTED_COLUMNS}
    summary = ImportSummary()
    seen_admission_numbers = set()
    class_level_lookup = {c.name.lower(): c for c in ClassLevel.objects.for_school(school)}
    section_lookup = {
        (s.class_level_id, s.name.lower()): s for s in Section.objects.for_school(school).select_related('class_level')
    }
    existing_admission_numbers = set(Student.objects.for_school(school).values_list('admission_number', flat=True))

    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = _row_values(row, column_index)
        if all(value in (None, '') for value in values.values()):
            continue

        summary.total_rows += 1
        errors = []

        admission_number = str(values.get('admission_number') or '').strip()
        first_name = str(values.get('first_name') or '').strip()
        last_name = str(values.get('last_name') or '').strip()
        class_level_name = str(values.get('class_level') or '').strip()
        section_name = str(values.get('section') or '').strip()

        if not admission_number:
            errors.append('Admission number is required.')
        if not first_name:
            errors.append('First name is required.')
        if not last_name:
            errors.append('Last name is required.')
        if not class_level_name:
            errors.append('Class is required.')

        class_level = class_level_lookup.get(class_level_name.lower()) if class_level_name else None
        if class_level_name and class_level is None:
            errors.append(f"Class '{class_level_name}' does not exist for this school.")

        section = None
        if section_name and class_level:
            section = section_lookup.get((class_level.id, section_name.lower()))
            if section is None:
                errors.append(f"Section '{section_name}' does not exist for class '{class_level_name}'.")

        if admission_number:
            if admission_number in existing_admission_numbers:
                errors.append('Admission number already exists for this school.')
            elif admission_number in seen_admission_numbers:
                errors.append('Duplicate admission number within this file.')

        if errors:
            summary.error_count += 1
            summary.row_results.append(RowResult(row_number, admission_number, 'error', errors))
            continue

        try:
            with transaction.atomic():
                Student.objects.create(
                    school=school,
                    admission_number=admission_number,
                    first_name=first_name,
                    last_name=last_name,
                    date_of_birth=values.get('date_of_birth') or None,
                    gender=str(values.get('gender') or '').strip().upper()[:10],
                    guardian_name=str(values.get('guardian_name') or '').strip(),
                    guardian_phone=str(values.get('guardian_phone') or '').strip(),
                    email=str(values.get('email') or '').strip(),
                    academic_year=academic_year,
                    class_level=class_level,
                    section=section,
                )
        except IntegrityError as exc:
            summary.error_count += 1
            summary.row_results.append(RowResult(row_number, admission_number, 'error', [str(exc)]))
            continue

        seen_admission_numbers.add(admission_number)
        summary.created_count += 1
        summary.row_results.append(RowResult(row_number, admission_number, 'created', []))

    return summary


def build_template_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Students'
    sheet.append(EXPECTED_COLUMNS)
    return workbook
